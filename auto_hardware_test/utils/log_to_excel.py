#!/usr/bin/env python3
# -*- coding:utf-8 -*-
import glob
import os
import configparser
import json
import re
import openpyxl

class Config:
    # 通过配置文件参数key，获取value
    def read_config(config_path, node, key):
        # 创建配置对象
        config = configparser.ConfigParser()
        # 读取ini文件
        config.read(config_path, encoding='utf-8')  # python3,若有中文，需要使用utf-8
        value = config.get(node, key)
        # print(value)
        return value

    # 回写配置到配置文件
    def write_config(config_path, node, key, value):
        # 回写当前镜像id，快照id到image.ini配置文件
        config = configparser.ConfigParser()
        config.read(config_path, encoding='utf-8')
        # a = config.add_section("image_id") #添加一个key
        config.set(node, key, str(value))
        # print(value)
        config.write(open(config_path, 'r+'))

    # 读取config文件节点
    def get_conf_sections(config_path):
        # 创建配置对象
        config = configparser.ConfigParser()
        # 读取ini文件
        config.read(config_path, encoding='utf-8')  # python3,若有中文，需要使用utf-8
        sections = config.sections()
        return sections

class Json:
    # 用来存储数据
    dict = {}

    # 获取json里面数据
    def get_json_data(json_path, key, value):
        # 定义为只读模型，并定义名称为f
        with open(json_path, 'rb') as f:
            # 加载json文件中的内容给params
            params = json.load(f)
            # 修改内容
            params[key] = value
            # print("params", params)
            # 将修改后的内容保存在dict中
            dict = params
        # 关闭json读模式
        f.close()
        # 返回dict字典内容
        return dict

    # 写入json文件
    def write_json_data(json_path, dict):
        # 定义为写模式，名称定义为r
        with open(json_path, 'w') as r:
            # 将dict写入名称为r的文件中
            json.dump(dict, r)
        # 关闭json写模式
        r.close()

    # 用来修改decom.json服务器IP
    def alter_serverIP(json_path, serverIP):
        # 定义为只读模型，并定义名称为f
        with open(json_path, 'rb') as f:
            # 加载json文件中的内容给load_dicts
            load_dicts = json.load(f)
            # print(load_dicts)
            for addr in load_dicts["service_addrs"]:
                if addr["dsc"].endswith(":7730"):
                    addr["dsc"] = serverIP + ":7730"
                if addr["dsc"].endswith(":7740"):
                    addr["dsc"] = serverIP + ":7740"
                if addr["dsc"].endswith(":7750"):
                    addr["dsc"] = serverIP + ":7750"
            # print(json.dumps(load_dicts, indent=4))
        f.close()
        return load_dicts

import csv
class Excel:

    # 使用openpyxl读excel：
    def read_excel_xlsx(path, sheetname):
        workbook = openpyxl.load_workbook(path)
        # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
        sheet = workbook[sheetname]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()

    # 通过测试用例名称使用openpyxl读excel下row：
    def read_excel_row(path, sheetname , minrow, maxrow, column, casename):
        workbook = openpyxl.load_workbook(path)
        # 取第一张表
        # sheetnames = workbook.get_sheet_names()  # 获取所有表格名
        # sheet = workbook.get_sheet_by_name(sheetname)  # 通过表名获取表格，最新python已不支持
        sheet = workbook[sheetname]

        # sheet = excel.active  #获取默认sheet
        # print(sheet.title)  # 输出表名
        for row in range(minrow, maxrow):
            test = sheet.cell(row=row, column=column)
            print(test.value)
            if test.value == casename:
                return row


    # 使用openpyxl对Excel内容追加写入
    def write_excel_xlsx(path, sheetname, row, column, value):
        workbook = openpyxl.load_workbook(path)
        # print(data.get_named_ranges())  # 输出工作页索引范围
        # print(data.get_sheet_names())  # 输出所有工作页的名称
        # 取第一张表
        # sheetnames = workbook.get_sheet_names()   # 获取所有表格名
        # sheet = workbook.get_sheet_by_name(sheet_name)  # 通过表名获取表格，最新python已不支持
        sheet = workbook[sheetname]

        # table = excel.active  #获取默认sheet
        # print(table.title)  # 输出表名
        sheet.cell(row, column, value)
        workbook.save(path)
        print("xlsx格式表格写入数据成功！行数：", row, ", 列数：", column, ", 值：", value)


class Log:

    # 获取后缀为xx的文件名，并存入列表
    # def get_filename_list(filepath, file_postfix):
    #     list_file = []
    #     for root, dirs, files in os.walk(filepath):
    #         for file in files:
    #             # 判断后缀为txt的
    #             if os.path.splitext(file)[1] == file_postfix:
    #                 list_file.append(os.path.splitext(file)[0])
    #         print(list_file)
    #     return list_file

    def get_filename_list(file_path_key):
        temp = []
        # 获得指定目录下的所有文件; 使用相对路径：
        file_name_list = glob.glob(r"" + file_path_key)
        for fn in range(0, len(file_name_list)):
            str = file_name_list[fn].replace("\\", "/")
            temp.append(str)
        print("列表展示",temp)
        return temp


    # 读取文件内容并返回结果
    def readFile(filename):
        data = []
        try:
            # 打开文件
            fp = open(filename, "r")
            print('%s 文件打开成功' % filename)
            for line in fp.readlines():
                '''
                当你读取文件数据时会经常遇见一种问题，
                    那就是每行数据末尾都会多个换行符‘\n’，
                    所以我们需要先把它们去掉
                '''
                line = line.replace('\n', '')
                # 或者line=line.strip('\n')
                # 但是这种只能去掉两头的，可以根据情况选择使用哪一种

                line = line.split(',')
                # 以逗号为分隔符把数据转化为列表

                data.append(line)
            fp.close()
            print("文件内容为：", data)
            return data
        except IOError:
            print("文件打开失败，%s文件不存在" % filename)

    # 寫入文件內容
    def write_file(file_name, content):
        # coding=UTF-8
        with open(file_name, 'w', encoding="UTF-8") as file_object:
            file_object.write(content)
            file_object.close()

    # 查找文件中指定内容
    def get_value(file_path, regex, key_word: str):
        with open(file_path, encoding="utf-8") as f:
            for index, line in enumerate(f.readlines()):
                if key_word in line:
                    # 利用正则表达式提取所需字符串
                    str = re.findall(regex, line) # re.findall输出['1122']#?控制只匹配0或1个,所以只会输出和最近的b之间的匹配情况
                    # print("findall--------", str)
                    result = re.sub(r'[\[\'\]]*', '', str.__str__())  # 前面是正则表达式，匹配多种字符（串）；去掉正则提取后的括号
                    message = f"{key_word}在文件的第{index + 1}行" + ", 值：" + result
                    break
                else:
                    result = ''
                    message = f"未在文件中发现{key_word}"
        return result

    # 查找文件中指定内容
    def get_value_list(file_path, regex, key_word: str):
        temp = []
        with open(file_path, encoding="utf-8") as f:
            for index, line in enumerate(f.readlines()):
                if key_word in line:
                    # 利用正则表达式提取所需字符串
                    str = re.findall(regex, line) #re.findall输出['1122']#?控制只匹配0或1个,所以只会输出和最近的b之间的匹配情况
                    print("str", str.__str__())
                    result = re.sub(r'[\[\'\]]*', '', str.__str__())  # 前面是正则表达式，匹配多种字符（串）；去掉正则提取后的括号
                    temp.append(result)
                else:
                    result = ''
                    message = f"未在文件中发现{key_word}"
        return temp

    # 查找文件中指定内容列表
    def get_key_line(file_path, key_word: str):
        temp = []
        with open(file_path) as f:
            for index, line in enumerate(f.readlines()):
                if key_word in line:
                    # 保存關鍵字所在行
                    result = re.sub(r'\'', '', line.__str__())  # 前面是正则表达式，去掉正则提取后的引號
                    temp.append(result)
                else:
                    result = ''
                    message = f"未在文件中发现{key_word}"
        return temp

    # 获取文件指定关键词的最后一行，并存入新文件
    # linux使用grep
    # def get_last_row(logfile, key, newfile):
    #     cmd = "grep "+key+" "+logfile+" | tail -1 > "+newfile
    #     print(cmd)
    #     os.system(cmd)
    #     #getcwd = os.getcwd()

    # 獲取正序最後文件名的最後一行，寫入到新文件
    def get_last_line(file_path_key, context_key_word, new_file):
        file_path_list = Log.get_filename_list(file_path_key)
        # 獲取正序最後一個文件路徑
        file_path = file_path_list[len(file_path_list) - 1]
        print("file_path--", file_path)
        line_list = Log.get_key_line(file_path, context_key_word)
        # 獲取文件最後一行
        last_line = line_list[len(line_list)-1]
        # 寫入新的文件
        Log.write_file(new_file, last_line)
        return last_line


    # 根据文件旧字符串修改为新字符串，保存到文件
    def alter(file, old_str, new_str):
        file_data = ""
        with open(file, "r", encoding="utf-8") as f:
            for line in f:
                if old_str in line:
                    line = line.replace(old_str, new_str)
                file_data += line
                # print(file_data)
        with open(file, "w", encoding="utf-8") as f:
            f.write(file_data)
        f.close()


    # 格式化文件路径，分成文件路径和文件名
    def formatFilePath(filePath):
        fpath = os.path.dirname(filePath) + '\\'
        filename = os.path.basename(filePath)
        return fpath, filename


class LogToExcel:
    # 读取fio日志，获取测试结果
    def read_fio_log(file):
        try:
            regex_fiotest = r"(.+?): \(g=0\):"
            fiotest = Log.get_value(file, regex_fiotest, "fiotest")
            print("fiotest:", fiotest)

            regex = r"IOPS=(.+?), BW"
            IOPS = Log.get_value(file, regex, "IOPS=")
            print("IOPS:", IOPS)

            regex_bw = r"BW=.+?\((.+?)MB/s\)"
            BW = Log.get_value(file, regex_bw, "BW=")
            if BW.__eq__(""):
                regex_bw = r"BW=.+?\((.+?GB/s)\)"
                BW = Log.get_value(file, regex_bw, "BW=")
            if BW.__eq__(""):
                regex_bw = r"BW=.+?\((.+?kB/s)\)"
                BW = Log.get_value(file, regex_bw, "BW=")
            print("BW:", BW)

            # 正则匹配为xx的数据，如果没匹配到usec，则匹配nsec，不为空则进制到usec并保留2位小数
            regex_lat = r"slat.+?avg=(.+?), stdev"  # lat.+?avg=(.+?), stdev
            slat_avg = Log.get_value(file, regex_lat, "slat (usec):")
            if slat_avg.__eq__(""):
                slat_avg = Log.get_value(file, regex_lat, "slat (nsec):")
                if slat_avg != "":
                    slat_avg = round(int(float(slat_avg)) / 1000, 2)
            # print("slat_avg:", slat_avg)

            # 正则匹配为xx的数据，如果没匹配到usec，则匹配nsec，不为空则进制到usec并保留2位小数
            regex_lat = r"clat.+?avg=(.+?), stdev"  # lat.+?avg=(.+?), stdev
            clat_avg = Log.get_value(file, regex_lat, "clat (usec):")
            if clat_avg.__eq__(""):
                clat_avg = Log.get_value(file, regex_lat, "clat (nsec):")
                if clat_avg != "":
                    clat_avg = round(int(float(clat_avg)) / 1000, 2)
            # print("clat_avg:", clat_avg)

            # 正则匹配为xx的数据，如果没匹配到usec，则匹配nsec，不为空则进制到usec并保留2位小数
            regex_lat = r" lat.+?avg=(.+?), stdev"  # lat.+?avg=(.+?), stdev
            lat_avg = Log.get_value(file, regex_lat, " lat (usec):")
            if lat_avg.__eq__(""):
                lat_avg = Log.get_value(file, regex_lat, " lat (nsec):")
                if lat_avg != "":
                    lat_avg = round(int(float(lat_avg)) / 1000, 2)
            print("lat_avg:", lat_avg)

            regex_cpu = r"cpu          : (.+?), ctx"
            cpu = Log.get_value(file, regex_cpu, "cpu          :")
            print("cpu:", cpu)

            # 读取所有延迟区间，存入列表，并遍历换行后存入字符串
            regex = r"(lat \(.+?\)   : .+)"
            delay_distribution = list(filter(None, Log.get_value_list(file, regex, "lat")))  # filter()函数可以过滤空字符和None
            delay_str = ""
            for dalay in delay_distribution:
                delay_str += dalay+"\n"
            print("delay_str:", delay_str)

            return fiotest, IOPS, BW, lat_avg, cpu, delay_str
        except Exception as e:
            print(str(e))

    # 提取fio日志信息，通过表格名、测试用例编号回写到Excel
    def write_fio_Excel(excel, sheetname, file_path, file_path_key):
        # fio_excel = "log/服务器性能测试结果统计20201105.xlsx"
        # sheetname = "磁盘基准测试-游戏盘"
        #
        # log_path = "log/fio_log"
        # log_postfix = ".log"

        # 判断传入的文件名是否存在，不存在跳出，存在则继续
        list_log = Log.get_filename_list(file_path_key)
        print("list_log---", list_log)

        # 如果list_log不为空
        if len(list_log):
            # 遍历日志文件
            for file in list_log:
                fiotest, IOPS, BW, lat_avg, cpu, delay_str = LogToExcel.read_fio_log(file)
                # 根据用例case判断行号
                column = 1
                row = Excel.read_excel_row(excel, sheetname, 3, 73, column, fiotest)  # 磁盘使用量2, 7 # 写入对应多少行（包头不包尾）数据！！！这里结尾是因为加上表头有2行！

                # 通过fiotest行号写入对应行
                Excel.write_excel_xlsx(excel, sheetname, row, 5, IOPS)  # 读IOPS
                Excel.write_excel_xlsx(excel, sheetname, row, 6, BW)  # 读BW
                Excel.write_excel_xlsx(excel, sheetname, row, 7, lat_avg)  # 读lat_avg
                Excel.write_excel_xlsx(excel, sheetname, row, 8, cpu)  # cpu使用率
                Excel.write_excel_xlsx(excel, sheetname, row, 9, delay_str)  # 延迟区间
        else:
            print("list_log为空，请检查fio日志是否存在！！！")



    # 读取镜像、游戏、回写服务日志，回写到Excel中
    def read_service_log(file_path, file_path_key, casename):  # , service_log, service_key, newfile, casename

        # 判断传入的文件名是否存在，不存在跳出，存在则继续
        list_log = Log.get_filename_list(file_path_key)
        filelist = []
        for log in list_log:
            strlist = re.split('[_/]', log)  # 一次指定多个分隔符可以用re模块
            filelist.extend(strlist)  # 分割并将字符串数组追加到文件名列表
        if casename not in filelist:
            print(casename, "日志文件不存在！")
            return 0
        print("strlist---", strlist)

        try:
            # 读取数据服务日志
            if casename.__eq__("dataservice"):
                context_key_word = "[gameservice_dump]|total"
                total_newfile = file_path + "data_total_row"
                total_regex = r"total:(.+?), inqueue"
                total_key = r"total"

                lay_key = "[gameservice_dump]|READ DELAY"
                lay_newfile = file_path + "data_lay_row"

                lay_regex = "\[.+?,(.+?)\):.+?;"  # 数据服务日志[**)的，最新区间
                lay_value_regex = "\[.+?\):\((.+?)\);"  # \[.+?\):(.+?);
                # 获取游戏服务总请求行
                data_total_row = Log.get_last_line(file_path_key, context_key_word, total_newfile)
                print(data_total_row)
                # 获取游戏服务总请求数

                data_total = Log.get_value(total_newfile, total_regex, total_key)
                print("data_total:", data_total)

                # 获取游戏服务总请求行延时分布
                data_lay_row = Log.get_last_line(file_path_key, lay_key, lay_newfile)
                print("data_lay_row:", data_lay_row)

                # # 拿到延迟区间存入list；此处无用
                # content = Log.readFile(lay_newfile)

                data_key_list = re.findall(lay_regex, data_lay_row)
                if len(data_key_list) == 0:
                    lay_regex = "\(.+?,(.+?)\]:.+?;"  # 数据服务日志[**)的，最新区间
                    data_key_list = re.findall(lay_regex, data_lay_row)
                print("key_list----", data_key_list.__str__())

                # 拿到延迟区间的值存入list
                data_value_list = re.findall(lay_value_regex, data_lay_row)
                if len(data_value_list) == 0:
                    regex_lay_value = "\(.+?\]:(.+?);"   # 数据服务日志[**)的，最新区间
                    data_value_list = re.findall(regex_lay_value, data_lay_row)
                print("value_list----", data_value_list.__str__())

                # 把延迟区间和值存入字典
                """ zip打包用法,同时遍历两个list """
                data_lay_dict = {}
                for k, v in zip(data_key_list, data_value_list):
                    if k not in data_lay_dict.keys():
                        data_lay_dict[k] = v  # 将键值对对应起来
                print("data_lay_dict:", data_lay_dict)

                return data_total, data_lay_dict

            # 读取镜像日志信息
            if casename.__eq__("diskservice"):
                context_key_word = "[diskservice_dump]|total"
                new_file = file_path + "disk_total_row"
                lay_key = "READ DELAY"
                lay_new_file = file_path + "disk_lay_row"

                # 获取镜像服务总请求行
                disk_total_row = Log.get_last_line(file_path_key, context_key_word, new_file)
                print(disk_total_row)
                # 获取镜像服务总请求数
                regex_total = r"total:(.+?), inqueue"
                key_total = r"total"
                disk_total = Log.get_value(new_file, regex_total, key_total)
                print("disk_total:", disk_total)
                # 获取镜像服务总请求延时分布

                disk_lay_row = Log.get_last_line(file_path_key, lay_key, lay_new_file)
                print("disk_lay_row:", disk_lay_row)

                # # 拿到延迟区间存入list；此处无用
                # content = Log.readFile(lay_new_file)

                regex_lay = "\(.+?,(.+?)]:.+?;"  # 数据服务日志[**)的，最新区间
                disk_key_list = re.findall(regex_lay, disk_lay_row)
                if len(disk_key_list) == 0:
                    regex_lay = "\[.+?,(.+?)\):.+?;"  # 数据服务日志[**)的，最新区间
                    disk_key_list = re.findall(regex_lay, disk_lay_row)
                print("key_list----", disk_key_list.__str__())

                # 拿到延迟区间的值存入list
                regex_lay_value = "\(.+?]:(.+?);"  # 数据服务日志[**)的，最新区间
                disk_value_list = re.findall(regex_lay_value, disk_lay_row)
                if len(disk_value_list) == 0:
                    regex_lay_value = "\[.+?\):\((.+?)\);"   # 数据服务日志[**)的，最新区间
                    disk_value_list = re.findall(regex_lay_value, disk_lay_row)
                print("value_list----", disk_value_list.__str__())

                # 把延迟区间和值存入字典
                """ zip打包用法,同时遍历两个list """
                disk_lay_dict = {}
                for k,v in zip(disk_key_list,disk_value_list):
                    if k not in disk_lay_dict.keys():
                        disk_lay_dict[k] = v  # 将键值对对应起来
                print("disk_lay_dict:", disk_lay_dict)

                return disk_total, disk_lay_dict

            # 讀取回寫日誌信息
            if casename.__eq__("writeback"):
                context_key_word = "[writeback_dump]|[write]"
                new_file = file_path + "writeback_total_row"
                key_total = r"total"

                read_regex_total = r"\[read] total/left:(.+?)/.+?, iops"
                read_lay_key = "[writeback_dump]|READ DELAY"
                read_lay_newfile = file_path + "read_lay_row"


                write_regex_total = r"\[write] total/left:(.+?)/.+?, iops"
                write_lay_key = "[writeback_dump]|WRITE DELAY"
                write_lay_newfile = file_path + "write_lay_row"

                # 获取回寫服务总请求行
                writeback_total_row = Log.get_last_line(file_path_key, context_key_word, new_file)
                print(writeback_total_row)

                regex_total_list = [read_regex_total, write_regex_total]
                lay_key_list = [read_lay_key, write_lay_key]
                lay_newfile_list = [read_lay_newfile, write_lay_newfile]
                result = []
                # 使用zip遍歷多個列表
                for regex_total, lay_key, lay_newfile in zip(regex_total_list, lay_key_list, lay_newfile_list):
                    # 获取回寫服务讀/寫总请求数
                    total = Log.get_value(new_file, regex_total, key_total)
                    print("total:", total)

                    # 获取回寫服务总请求延时分布
                    lay_row = Log.get_last_line(file_path_key, lay_key, lay_newfile)
                    print("lay_row:", lay_row)

                    # 拿到延迟区间存入list
                    regex_lay = "\(.+?,(.+?)]:.+?;"  # 数据服务日志[**)的，最新区间
                    key_list = re.findall(regex_lay, lay_row)
                    print("key_list----", key_list.__str__())

                    # 拿到延迟区间的值存入list
                    regex_lay_value = "\(.+?]:(.+?);"  # 数据服务日志[**)的，最新区间
                    value_list = re.findall(regex_lay_value, lay_row)
                    print("value_list----", value_list.__str__())

                    # 把延迟区间和值存入字典
                    """ zip打包用法,同时遍历两个list """
                    lay_dict = {}
                    for k,v in zip(key_list, value_list):
                        if k not in lay_dict.keys():
                            lay_dict[k] = v  # 将键值对对应起来
                    print("lay_dict:", lay_dict)
                    result.append(total)
                    result.append(lay_dict)
                print("result---", result)
                return result
        except Exception as e:
            print(str(e))



    # 回写镜像、游戏、回写服务日志延迟区间及值到Excel
    def write_service_Excel(excel, sheetname, file_path, file_path_key, casename):
        # service_excel = "log/服务器性能测试结果统计20201105.xlsx"
        # sheetname = "服务端压测结果统计"

        # 判断需要写入的服务到Excel
        if casename.__eq__("dataservice"):
            # 遍历日志文件
            data_total, data_lay_dict = LogToExcel.read_service_log(file_path, file_path_key, casename)
            if data_total and data_lay_dict:
                # 根据用例case判断行号
                column = 3  # 服务名称所在列
                row = Excel.read_excel_row(excel, sheetname, 3, 7, column, casename)  # 磁盘使用量2, 7 # 写入对应多少行（包头不包尾）数据！！！
                print("row:", row)
                column = 4  # total_request对应列
                print("column:", column)

                # 通过casename行号写入对应行
                Excel.write_excel_xlsx(excel, sheetname, row, column, data_total)

                print("data_lay_dict:::", data_lay_dict)
                # 写入游戏服务表头
                row = row - 1
                column = column + 2
                for data_title in data_lay_dict.keys():
                    Excel.write_excel_xlsx(excel, sheetname, row, column, data_title)
                    column += 1
                # 写入游戏服务延迟区间对应值
                row = row + 1  # 为了适应表头
                column = 6  # 延迟区间首列
                for data_value in data_lay_dict.values():
                    Excel.write_excel_xlsx(excel, sheetname, row, column, data_value)
                    column += 1
                print("dataservice_write_Excel")
            else:
                print("dataservice延迟信息为空，无法写入Excel！")

        if casename.__eq__("diskservice"):
            # 遍历日志文件
            disk_total, disk_lay_dict = LogToExcel.read_service_log(file_path, file_path_key, casename)
            if disk_total and disk_lay_dict:
                # 根据用例case判断行号
                column = 3  # 服务名称所在列
                row = Excel.read_excel_row(excel, sheetname, 3, 7, column, casename)  # 磁盘使用量2, 7 # 写入对应多少行（包头不包尾）数据！！！
                print("row:", row)
                column = 4  # total_request对应列
                print("column:", column)

                # 通过casename行号写入对应行
                Excel.write_excel_xlsx(excel, sheetname, row, column, disk_total)

                print("disk_lay_dict:::", disk_lay_dict)
                # 写入表头
                # column = column + 2
                # row += 1
                # for disk_title in disk_lay_dict.keys():
                #     Excel.write_excel_xlsx(service_excel, sheetname, row, column, disk_title)
                #     column += 1
                # 写入镜像服务延迟区间对应值
                column = 6  # 延迟区间首列
                # row -= 1  # 为了适应表头
                for disk_value in disk_lay_dict.values():
                    Excel.write_excel_xlsx(excel, sheetname, row, column, disk_value)
                    column += 1

                print("diskservice_write_Excel")
            else:
                print("diskservice延迟信息为空，无法写入Excel！")


        if casename.__eq__("writeback"):
            # 遍历日志文件
            result = LogToExcel.read_service_log(file_path, file_path_key, casename)

            # 根据用例case判断行号
            column = 3  # 服务名称所在列
            row = Excel.read_excel_row(excel, sheetname, 3, 7, column, casename)  # 磁盘使用量2, 7 # 写入对应多少行（包头不包尾）数据！！！
            print("row:", row)
            if result:
                read_list = [result[0], result[1]]
                write_list = [result[2], result[3]]
                writeback_list = [read_list, write_list]
                for i in range(len(writeback_list)):
                    column = 4  # total_request对应列
                    # 通过casename行号写入对应行
                    Excel.write_excel_xlsx(excel, sheetname, row, column, writeback_list[i][0])
                    # column += 1
                    print("行列：", column, row, "total:::", writeback_list[i][0], "data_lay_dict:::", writeback_list[i][1])
                    # 写入表头
                    # for disk_title in disk_lay_dict.keys():
                    #     Excel.write_excel_xlsx(service_excel, sheetname, row, column, disk_title)
                    #     column += 1
                    # 写入镜像服务延迟区间对应值
                    column = 6  # 延迟区间首列
                    for disk_value in writeback_list[i][1].values():
                        Excel.write_excel_xlsx(excel, sheetname, row, column, disk_value)
                        column += 1
                    row += 1  # 回寫中寫服務日誌延遲區間寫入下一行
                print("writeback_write_Excel")
            else:
                print("writeback延迟信息为空，无法写入Excel！")


    # 读取客户端压测count日志，回写到Excel中
    def read_count_log(file_path_key, file_path, ip):
        # 判断传入的文件名是否存在，不存在跳出，存在则继续
        list_log = Log.get_filename_list(file_path_key)
        filelist = []
        for log in list_log:
            strlist = re.split('[_/]', log)  # 一次指定多个分隔符可以用re模块
            filelist.extend(strlist)  # 分割并将字符串数组追加到文件名列表
        if ip not in filelist:
            print(ip, "_count日志文件不存在！")
            return 0

        try:
            # file_path_key = file_path + ip + "_count_*.log"
            context_key_word = "total_request"
            new_file = file_path + ip +"_count_total_row"
            total_filepath = file_path + ip + "_count_total_row"
            total_regex = r"total_request:(.+?); total_distribution"
            total_key = r"total_request"

            total_distribution_regex = "total_distribution:(.+)"
            total_distribution_key = "total_distribution"

            lay_regex = "(.+?):.+?;"
            lay_value_regex = ".+?:(.+?);"
            # 获取count统计总请求行
            # total_request_row = Log.get_last_row(count_log, count_key, count_newfile)
            # count_key = ""
            total_request_row = Log.get_last_line(file_path_key, context_key_word, new_file)
            print("total_request_row:", total_request_row)

            # 获取count统计总请求数
            total_request = Log.get_value(total_filepath, total_regex, total_key)
            print("total_request:", total_request)

            # 获取游戏服务总请求行延时分布
            total_distribution = Log.get_value(total_filepath, total_distribution_regex, total_distribution_key).replace(" ", "")
            print("total_distribution:", total_distribution)

            # 拿到延迟区间存入list
            key_list = re.findall(lay_regex, total_distribution)
            print("key_list----", key_list.__str__())

            # 拿到延迟区间的值存入list
            value_list = re.findall(lay_value_regex, total_distribution)
            print("value_list----", value_list.__str__())

            # 把延迟区间和值存入字典
            """ zip打包用法,同时遍历两个list """
            lay_dict = {}
            for k, v in zip(key_list, value_list):
                if k not in lay_dict.keys():
                    lay_dict[k] = v  # 将键值对对应起来
            print("data_lay_dict:", lay_dict)

            return total_request, lay_dict
        except Exception as e:
            print(str(e))


    def write_count_Excel(service_excel, sheetname, file_path, file_path_key, ip):
        # service_excel = "log/服务器性能测试结果统计20201105.xlsx"
        # sheetname = "服务端压测结果统计"

        # 判断需要写入的服务到Excel
        # 遍历日志文件
        total_request, lay_dict = LogToExcel.read_count_log(file_path_key, file_path, ip)
        # 根据用例case判断行号
        column = 3  # ip所在列
        row = Excel.read_excel_row(service_excel, sheetname, 3, 7, column, ip)  # 磁盘使用量2, 7 # 写入对应多少行（包头不包尾）数据！！！
        print("row:", row)
        column = 5  # total_request对应列
        print("column:", column)

        # 通过hostname行号写入对应行
        Excel.write_excel_xlsx(service_excel, sheetname, row, column, total_request)

        print("lay_dict:::", lay_dict)
        # 写入count表头
        row_title = 2
        column = column + 2
        for count_title in lay_dict.keys():
            Excel.write_excel_xlsx(service_excel, sheetname, row_title, column, count_title)
            column += 1

        # 写入count延迟区间对应值
        column = 7
        for count_value in lay_dict.values():
            Excel.write_excel_xlsx(service_excel, sheetname, row, column, count_value)
            column += 1


if __name__ == '__main__':

    excel = "../服务器性能测试结果统计.xlsx"

    # 读取磁盘基准测试fio日志，并将延迟等回写到Excel
    sheetname = "磁盘基准测试-游戏盘"
    log_path = "../log/"
    log_path_key = log_path + "fiotest*.log"
    LogToExcel.write_fio_Excel(excel, sheetname, log_path, log_path_key)
    #
    # 读取磁盘基准测试fio日志，并将延迟等回写到Excel
    # sheetname = "磁盘基准测试-镜像盘"
    # log_path = "../log/fio_log"
    # log_postfix = ".log"
    # LogToExcel.write_fio_Excel(excel, sheetname, log_path, log_postfix)
    #
    # 读取磁盘基准测试fio日志，并将延迟等回写到Excel
    # sheetname = "磁盘基准测试-回写盘"
    # log_path = "../log/"
    # log_postfix = ".log"
    # LogToExcel.write_fio_Excel(excel, sheetname, log_path, log_postfix)


    # # # 读取游戏服务日志并写入Excel
    #这儿
    # sheetname = "capture压测服务端统计"
    # file_path = "../log/"
    # file_path_key = file_path + "dataservice_*.log"
    # casename="dataservice"
    # LogToExcel.write_service_Excel(excel, sheetname, file_path, file_path_key, casename)
    #
    #
    # # 读取镜像服务日志并写入Excel
    # file_path_key = file_path + "diskservice_*.log"
    # casename = "diskservice"
    # LogToExcel.write_service_Excel(excel, sheetname, file_path, file_path_key, casename)
    # #
    # # # 读取回写服务日志并写入Excel
    # # file_path_key = file_path + "writeback_*.log"
    # # casename = "writeback"
    # # # LogToExcel.read_service_log(file_path, file_path_key, casename)
    # # LogToExcel.write_service_Excel(excel, sheetname, file_path, file_path_key, casename)
    # #
    #
    # # 读取客户端count日志并提取延迟区间写入Excel
    # sheetname = "capture压测客户端统计"
    # config_path = "../config.conf"
    # file_path = "../log/"
    #
    # # 读取count并写入Excel
    # sections = Config.get_conf_sections(config_path)
    # for i in range(2, len(sections)):
    #     section = sections[i]
    #     ip = Config.read_config(config_path, section, "host")
    #     file_path_key = file_path + ip.__str__() + "_count_*.log"
    #     print("ip---", ip)
    #     LogToExcel.write_count_Excel(excel, sheetname, file_path, file_path_key, ip)
